import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime


class ReadXLSX:

    def __init__(self, path):
        self.path = path
        self.workbook = openpyxl.load_workbook(path)
        self.rows = self.__readFile()
        self.uniqeProductNames = self.__getUniqeProductsName()
        self.uniqeProducts = self.__getUniqeProductList()
        self.dateInfos = self.__getDate()
        self.df = self.__getDf()
        # self.__writeXLSX()

    def __readFile(self):
        wb_obj = self.workbook
        sheet_obj = wb_obj.active
        global data

        if len(wb_obj.sheetnames) == 1:
            data = wb_obj.create_sheet("Data")
        else:
            wb_obj.remove(wb_obj["Data"])
            data = wb_obj.create_sheet("Data")

        startRow = 5
        startCol = 1
        finishRow = sheet_obj.max_row - 1
        finishCol = sheet_obj.max_column

        rows = []

        for x in range(startRow, finishRow + 1):
            col = []
            for y in range(startCol, finishCol):
                cell_obj = sheet_obj.cell(row=x, column=y)
                col.append(cell_obj.value)

            rows.append(col)
        return rows

    def __getUniqeProductsName(self):
        uniqeProduct = []
        rows = self.rows
        urunler = []
        for urun in rows:
            urunler.append(urun[2])

        if urunler:
            for item in urunler:
                if item not in uniqeProduct:
                    uniqeProduct.append(item)
        else:
            return urunler

        return uniqeProduct

    def __getProduct(self, urunName):
        rows = self.rows
        satis = 0
        ortBirimFiyat = 0
        i = 0;
        toplamTutar = 0
        for x in rows:
            if x[2] == str(urunName):
                ortBirimFiyat += x[9]
                satis = satis + x[6]
                toplamTutar += x[10]
                i = i + 1
                urunIsmi = x[2]
                porsiyon = x[3]
                grup = x[4]
        ortBirimFiyat = ortBirimFiyat / i

        urun = [urunIsmi, porsiyon, grup, satis, ortBirimFiyat, toplamTutar]
        return urun

    def __getUniqeProductList(self):
        urunler = []
        for x in self.uniqeProductNames:
            urunler.append(self.__getProduct(x))
        return urunler

    def __writeXLSX(self):
        data.append(["Başlangıç Tarihi", "Bitiş Tarihi", "Gün Farkı"])
        data.append(self.__getDate())
        data.append(["Ürün Adı", "Porsiyon", "Kategori", "Satış Adeti", "Ortalama Birim Fiyat", "Toplam Tutar"])
        urunList = self.uniqeProducts
        for row in urunList:
            data.append(row)

        tab = Table(displayName="Table1", ref="A3:F" + str(len(urunList) + 1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        data.add_table(tab)
        self.workbook.save(self.path)

    def __getDate(self):
        wb_obj = self.workbook

        referenceSheet = wb_obj.active
        date = referenceSheet.cell(row=3, column=1)

        dates = date.value.split(' - ')

        d1 = datetime.strptime(dates[0], '%d.%m.%Y %H:%M:%S')
        d2 = datetime.strptime(dates[1], '%d.%m.%Y %H:%M:%S')
        different = abs((d2 - d1).days)
        d1 = d1.strftime("%d.%m.%Y")
        d2 = d2.strftime("%d.%m.%Y")
        dateObj = [str(d1), str(d2), str(different)]

        return dateObj

    def __getDf(self):
        df = pd.DataFrame(self.__getUniqeProductList(),
                          columns=["ÜrünAdı", "Porsiyon", "Kategori", "Satış Adeti", "Ortalama Birim Fiyat",
                                   "Toplam Tutar"])
        return df

    def getSumByCategory(self):
        df = self.df
        df2 = df.groupby("Kategori").sum()

        # print(df2[["Satış Adeti", "Toplam Tutar"]])
        return df2[["Satış Adeti", "Toplam Tutar"]]

    def getPercentByCategory(self):
        df = self.getSumByCategory()
        df['Satış Adeti'] = (df['Satış Adeti'] / df['Satış Adeti'].sum()) * 100
        df['Toplam Tutar'] = (df['Toplam Tutar'] / df['Toplam Tutar'].sum()) * 100

        return df
